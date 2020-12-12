#usr/bin/python
#########################
### Libraries Used
#########################
import time
import sys
import re
import os
import StringIO
import xlrd
import xlwt

from openpyxl import load_workbook
wb = load_workbook('./test.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
market='Bharti-Chennai'
print ("market",market)
hostip="0.0.0.0"
x=1
while (x < 3):
	print (type(market))
	print (type(sheet.cell(row=x, column=1).value))
	if market == sheet.cell(row=x, column=1).value:
		hostip = sheet.cell(row=x, column=2).value
	
	print ("inside",sheet.cell(row=x, column=2).value)
	






def findIPfromMarket(market):from openpyxl import load_workbook
wb = load_workbook('./test.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet['B2'].value
	workbook = xlrd.open_workbook('SON_MARKET_LIST.xlsx')
	worksheet = workbook.sheet_by_name('Sheet1')
	x = 0
	for market in range (2):
		print(type(market))
		print(type(worksheet.cell(x, 0).value))
		if market == sheet.cell(row=x, column=0).value:
			hostip = sheet.cell(row=x, column=0).value
			break
		print("inside",sheet.cell(row=x, column=0).value)
		x = x + 1
findIPfromMarket(market)
print("hostip",hostip)


pip install openpyxl
print (type(x))

for x in range (1,3):
    print(type(market))
    print(type(sheet.cell(row=x, column=1).value))
    if market == sheet.cell(row=x, column=1).value:
        hostip = sheet.cell(row=x, column=2).value
        break
    print("inside",sheet.cell(row=x, column=2).value)
    x = x+1

